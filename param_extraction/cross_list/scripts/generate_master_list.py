# SPDX-FileCopyrightText: 2026 Contributors to the RISCV UnifiedDB <https://github.com/riscv/riscv-unified-db>
# SPDX-License-Identifier: BSD-3-Clause-Clear
"""Build the master parameter list: one row per distinct parameter concept.

  out/master_list.csv   -- the source of truth, one row per concept
  out/master_list.adoc  -- the same table, rendered for reading

Scope: the union of James Ball's list and ours, **minus** the entries UDB
already defines outright. Everything else is kept, including the entries that
only partially overlap a UDB parameter, because a partial overlap is not
coverage -- his register-level ``htinst`` against UDB's fourteen
``TINST_VALUE_ON_*`` is a granularity mismatch, not a duplicate. The
``udb_param`` column carries that overlap where it exists.

Design decisions worth knowing before reading the code:

* Rows are keyed by **concept**, not by name. Exactly one of our 38 names
  matches one of his, so a name-keyed union would double-count nearly every
  shared parameter. The concept join is the one already established in
  ``generate_asciidoc.py`` (normative tag first, then hand-verified concept
  matching), and this script reuses it rather than re-deriving it, so the
  master list cannot drift from the comparison document.

* A shared row can cite several of his names. 30 shared rows consume 35 of
  his entries because some of ours are umbrellas -- ``STANDARD_INTERRUPT_-
  SUPPORT`` covers four of his, ``DELEGATABLE_EXCEPTIONS`` two.

* ``confidence`` is deliberately emitted **empty**. It is a review column for
  humans to fill, not something to guess at; the scale is documented in the
  file header so entries stay comparable.

* Known defects are held in the DEFECTS table below rather than being written
  into prose, so each can be challenged on its own.
"""

import csv
import json
import re
import sys
from pathlib import Path

import yaml

sys.path.insert(0, str(Path(__file__).resolve().parent))

import generate_asciidoc as G

OUT_CSV = G.OUTDIR / "master_list.csv"
OUT_XLSX = G.OUTDIR / "master_list.xlsx"
OUT_ADOC = G.OUTDIR / "master_list.adoc"

CONFIDENCE_SCALE = "high | medium | low"

# Every field of his schema, carried through verbatim rather than folded into
# a derived value. ``james_domain`` normalises type/range for the like-for-like
# comparison against ours; these are the unprocessed originals beside it, so
# nothing of his is lost to that normalisation. Column name -> his field name.
JAMES_RAW_FIELDS = [
    ("james_section", "section"),
    ("james_chapter", "chapter"),
    ("james_granularity", "granularity"),
    ("james_reg_name", "reg-name"),
    ("james_field_name", "field-name"),
    ("james_long_name", "long-name"),
    ("james_description", "description"),
    ("james_type", "type"),
    ("james_width", "width"),
    ("james_range", "range"),
    ("james_array", "array"),
    ("james_func_of_field", "func-of-field-name"),
    ("james_note", "note"),
]

COLUMNS = [
    "id",
    "concept",
    "on_lists",
    "udb_param",
    "udb_relationship",
    "james_name",
    "james_domain",
    "james_source",
    "ishaan_name",
    "ishaan_domain",
    "ishaan_source",
    "ishaan_class",
    "match_basis",
    "review_status",
    "confidence",
    "notes_errors",
] + [c for c, _ in JAMES_RAW_FIELDS]


# One-line description of the choice each of our 38 parameters captures,
# written off the spec excerpt recorded in ours_canonical.json. Kept as a
# table so each label can be checked against its sentence.
CONCEPTS = {
    "CSR_STRONGLY_ORDERED":
        "Which CSRs, if any, the platform defines as strongly ordered",
    "CTRTARGET_MISP_IMPLEMENTED":
        "Whether ctrtarget.MISP is implemented, else read-only 0",
    "CTR_CCE_WIDTH":
        "How many exponent bits (0..4) are implemented in CCE",
    "CTR_CTRDATA_TYPE_IMPLEMENTED":
        "Which fields of ctrdata are implemented, else read-only 0",
    "CTR_CYCLE_COUNTING_SUPPORTED":
        "Whether ctrdata includes a cycle count since the prior CTR record",
    "DELEGATABLE_EXCEPTIONS":
        "Which medeleg/mideleg bits are delegatable (writable vs read-only 0)",
    "HGATP_PPN_LOWER_BITS_RO":
        "Whether hgatp.PPN[1:0] are read-only zero",
    "HINT_SXLEN_DEST_REG_BEHAVIOR":
        "Whether a HINT with XLEN < SXLEN overwrites dest-register bits "
        "SXLEN..XLEN",
    "HINT_XLEN_REDUCTION_BEHAVIOR":
        "Whether a HINT with XLEN < MXLEN overwrites dest-register bits "
        "MXLEN..XLEN",
    "HIP_BIT_WRITABLE":
        "Whether hip bit i is writable when sie bit i is read-only zero",
    "HPM_MISCONFIGURED_BEHAVIOR":
        "Whether a misconfigured HPM counter returns a constant value",
    "HPM_UNIMPLEMENTED_ACCESS_BEHAVIOR":
        "Whether accessing an unimplemented counter traps or returns a "
        "constant",
    "HTVAL_LEGAL_VALUES":
        "Which 2-bit-shifted guest physical addresses htval can hold",
    "MCYCLE_SHARED":
        "Whether mcycle (and mcountinhibit.CY) is shared between harts on a "
        "core",
    "MENVCFG_FIOM_READONLY":
        "Whether menvcfg.FIOM is read-only zero when S-mode is absent or "
        "satp.MODE is read-only zero",
    "MIP_WRITABLE_BITS":
        "Which mip bits are writable vs read-only",
    "MNEPC_INVALID_ADDRESS_CONVERSION":
        "Whether an invalid address is converted before being written to "
        "mnepc",
    "MSECCFG_SEED_BITS_RW":
        "Whether mseccfg [s,u]seed are writable or a read-only constant 0",
    "MSTATEEN_BIT63_TYPE":
        "Whether mstateen bit 63 is read-only zero (conditional on H absent "
        "and sstateen all read-only zero)",
    "MTVAL2_LEGAL_VALUES":
        "Which 2-bit-shifted guest physical addresses mtval2 can hold",
    "PMA_IDEMPOTENT_IMPLICIT_READ_SIZE":
        "The size of the naturally aligned power-of-2 region for idempotent "
        "implicit reads",
    "PMPADDR_WARL_MASK":
        "Which physical address bits are implemented in pmpaddr",
    "SCTRDEPTH_SUPPORTED_VALUES":
        "Which sctrdepth DEPTH values are supported",
    "SENVCFG_FIOM_ACCESS":
        "Whether senvcfg.FIOM is read-only zero when satp.MODE is read-only "
        "zero",
    "SEPC_INVALID_ADDR_BEHAVIOR":
        "Which invalid addresses sepc can hold, and whether writes are "
        "converted",
    "SIP_BITS_ACCESS":
        "Which sip bits are writable vs read-only",
    "SISELECT_WIDTH":
        "The supported value range of siselect (0..0xFFF minimum)",
    "SSTATUS_UBE_ACCESS":
        "Whether sstatus.UBE is read-only, mirroring S-mode endianness",
    "SSTATUS_UXL_ACCESS":
        "Whether sstatus.UXL is read-only, forcing UXLEN=SXLEN",
    "STANDARD_INTERRUPT_SUPPORT":
        "Which standard interrupt types (SEI/STI/SSI/LCOFI) are implemented, "
        "else pending/enable bits read-only 0",
    "STATEEN_IMPLICIT_UPDATE_EXCEPTION":
        "Whether an implicit state update blocked by a stateen CSR raises an "
        "illegal-instruction exception",
    "VSISELECT_WIDTH":
        "The supported value range of vsiselect (0..0xFFF minimum)",
    "VSSTATUS_UBE_PARAM":
        "Whether vsstatus.UBE is a read-only copy of hstatus.VSBE",
    "VSXL_RO_PARAM":
        "Whether hstatus.VSXL is read-only, forcing VSXLEN=HSXLEN",
    "VTW_VIRTINSTR_PARAM":
        "Whether WFI always raises a virtual-instruction exception in VS-mode "
        "when VTW=1",
    "WFI_TW_ALWAYS_ILLEGAL":
        "Whether WFI always raises an illegal-instruction exception below "
        "M-mode when TW=1",
    "XRET_CLEARS_LR_RESERVATION":
        "Whether xRET clears an outstanding LR reservation",
    "ZALASR_MISALIGNED_ATOMICITY_GRANULE":
        "Whether the misaligned atomicity granule PMA relaxes the Zalasr "
        "alignment requirement",
}


# Three of his entries define themselves purely as an AsciiDoc value table, so
# no prose line can be lifted as a label. Written off his own value tables.
JAMES_CONCEPTS = {
    "MSTATUS_FS_IMPRECISE":
        "Whether mstatus.FS goes Dirty only on real FP state change, or "
        "dirtiness is not tracked at all",
    "MSTATUS_VS_SW_DIRTY_UPDATE":
        "Whether a software write of Initial or Clean to mstatus.VS is forced "
        "to Dirty",
    "MTVAL2_TRAPVAL":
        "What mtval2 is written with on a guest page fault (zero vs the guest "
        "physical address)",
}


# Defects found while verifying, keyed by the entry they land on. Each is a
# factual claim about a file, not a judgement about the parameter.
DEFECTS = {
    "hgainp.MODE":
        "DEFECT (his file): hypervisor.yaml:106 declares reg-name `hgainp`, "
        "but its own impl-def is HGATP_MODE_WARL. `hgainp` is not a RISC-V "
        "CSR; typo for `hgatp`.",
    "MIPMPID":
        "DEFECT (his file): machine.yaml names the parameter MIPMPID but its "
        "impl-def is IMP_ID and the CSR is `mimpid`. Typo.",
    "VECTOR_LS_MISSALIGNED_EXCEPTION":
        "DEFECT (his file): spelled MISSALIGNED with a doubled S.",
    "VECTOR_LS_WHOLEREG_MISSALIGNED_EXCEPTION":
        "DEFECT (his file): spelled MISSALIGNED with a doubled S.",
    "PMA_MM_IFETCH":
        "DEFECT (his file): cites a tag whose sentence does not describe the "
        "parameter -- 'Main memory regions always support read and write' "
        "says nothing about instruction fetch.",
    "sctrdepth.DEPTH":
        "DEFECT (his file): the same field is recorded in both sections "
        "(SCTRDEPTH_DEPTH under parameter_definitions, sctrdepth.DEPTH under "
        "csr_definitions). The two carry different impl-defs so the split may "
        "be deliberate; worth confirming with him.",
    "VECTOR_LS_VSTART":
        "DEFECT (his file): named for vstart, but the description asks the "
        "overwrite-past-trap question, which duplicates his own "
        "VECTOR_LS_OVERWRITE_PAST_TRAP. Mapped to UDB LEGAL_VSTART on the "
        "name; by description it would map to UDB VECTOR_LOAD_PAST_TRAP. "
        "Needs his clarification before the row can be settled.",
    "vstvec.BASE":
        "OPEN (our bookkeeping): classified as a partial overlap citing "
        "VSTVEC_MODE_DIRECT/VECTORED, but those are MODE parameters and the "
        "stated reason is that UDB has no vstvec BASE-alignment parameter. "
        "Its sibling stvec.BASE is classified as having no UDB counterpart. "
        "Reads like it should be no-counterpart too; left unchanged pending "
        "review.",
}


# An adjudication reason is written to justify the UDB verdict, so for the
# partial-overlap rows it opens by comparing against UDB rather than by saying
# what the parameter is. Those openings are not usable as a concept label.
COMPARISON_LEAD = re.compile(
    r"^(UDB|Same|He |His |Adjacent|Related|Jordan|Verified|Duplicates)")


def scrub(v):
    """Collapse whitespace so no cell carries a newline into a spreadsheet."""
    return re.sub(r"\s+", " ", str(v or "")).strip()


def render_raw(v):
    """Render one of his field values without interpreting it."""
    if v is None:
        return ""
    if isinstance(v, list):
        return ", ".join(str(x) for x in v)
    return str(v)


def james_raw():
    """Every field of his entries, verbatim, keyed by the name we display.

    Read straight from his YAML rather than from ``load_james``, which keeps
    only the fields the comparison needed. ``width``, ``array``, ``note`` and
    ``func-of-field-name`` exist nowhere else in this pipeline.
    """
    out = {}
    for p in sorted(G.JDIR.glob("*.yaml")):
        doc = yaml.safe_load(p.read_text()) or {}
        chapter = doc.get("chapter_name") or p.stem
        for e in doc.get("parameter_definitions") or []:
            out[e["name"]] = dict(
                e, section="parameter_definitions", chapter=chapter,
                granularity="n/a")
        for e in doc.get("csr_definitions") or []:
            reg = e.get("reg-name") or e.get("reg-names")
            reg = reg if isinstance(reg, str) else "/".join(reg)
            fld = e.get("field-name")
            name = f"{reg}.{fld}" if fld else reg
            out[name] = dict(
                e, section="csr_definitions", chapter=chapter,
                granularity="field-level" if fld else "register-level")
            # reg-names (the four-register stateen entries) normalised so the
            # column is populated whichever spelling his file used
            out[name].setdefault("reg-name", reg)
    return out


def clean_description(text):
    """His descriptions often lead with an AsciiDoc value table.

    Take the first line of real prose and drop the markup, so a concept label
    never carries ``[%autowidth]`` or a ``!===`` row into the spreadsheet.
    """
    for line in (text or "").splitlines():
        line = line.strip()
        if not line or line.startswith(("[", "!", "|", "=")):
            continue
        # A wrapped continuation of a table row is not a sentence start.
        if not (line[0].isupper() or line[0] in "`_"):
            continue
        return re.sub(r"\s+", " ", line)
    return ""


def synth_concept(e):
    """Fallback label for an entry of his that carries no description."""
    if e["granularity"] == "field-level":
        return f"Legal values / access behaviour of `{e['name']}`"
    if e["granularity"] == "register-level":
        return (f"Legal values / access behaviour of `{e['name']}` "
                "(whole register)")
    return f"`{e['name']}` (no description in his file)"


def first_sentence(text):
    """Split a reason into its leading claim and the remaining evidence."""
    text = (text or "").strip()
    if not text:
        return "", ""
    for i in range(len(text) - 1):
        if text[i] == "." and (text[i + 1] == " " or i == len(text) - 1):
            return text[:i + 1].strip(), text[i + 1:].strip()
    return text, ""


def notes_for(bits):
    return "  ".join(b for b in bits if b)


def build():
    rules, by_tag = G.load_rules()
    ours, _meta = G.load_ours(by_tag, rules)
    james = G.load_james(rules)
    raw = james_raw()
    verdicts = G.ADJ["verdicts"]
    audit = {
        r["name"]: r
        for r in json.loads((G.BASE / "data/james_criteria_audit.json")
                            .read_text())["rows"]
    }

    j_by_tag, j_by_name = {}, {}
    for r in james:
        j_by_name[r["name"]] = r
        for t in r["tags"]:
            j_by_tag.setdefault(t, []).append(r)

    def j_cell(names, key):
        """Join a field across the several of his entries a row may cite."""
        vals = []
        for n in names:
            e = j_by_name.get(n)
            if not e:
                continue
            v = e[key] if key != "source" else "; ".join(e["impl_defs"])
            if v and v not in vals:
                vals.append(v)
        return " | ".join(vals)

    def j_raw(names):
        """His raw fields for the entries a row cites.

        A row can cite several of his entries -- our umbrella parameters map
        one-to-many onto his -- so values are joined with ' | ' in the order
        the names appear, and identical values are collapsed.
        """
        cells = {}
        for col, field in JAMES_RAW_FIELDS:
            vals = []
            for n in names:
                v = render_raw(raw.get(n, {}).get(field))
                if v and v not in vals:
                    vals.append(v)
            cells[col] = " | ".join(vals)
        return cells

    def j_notes(names):
        """Per-entry defects and criteria flags for his cited entries."""
        bits = []
        for n in names:
            if n in DEFECTS:
                bits.append(DEFECTS[n])
            a = audit.get(n)
            if a and not a.get("resolved"):
                bits.append(
                    f"His `{n}` cites impl-def "
                    f"{', '.join(a.get('tags') or [a.get('impl') or '?'])} "
                    "with no matching normative rule in the pinned manual, so "
                    "its spec text could not be checked.")
            for h in (a or {}).get("hits", []):
                bits.append(
                    f"Flagged by our INCLUSION_CRITERIA rule {h['rule']} "
                    f"on `{n}`: {h['why']}.")
        return bits

    rows = []

    # --------------------------------------------------- shared + ours-only
    matched, arguable, only_ours = [], [], []
    for r in sorted(ours, key=lambda x: x["name"]):
        tag = r["tags"][0] if r["tags"] else None
        hits = j_by_tag.get(tag, []) if tag else []
        if hits:
            matched.append((r, [h["name"] for h in hits], "tag", ""))
            continue
        if r["name"] in G.MANUAL_MATCHES:
            jn, _rel, why = G.MANUAL_MATCHES[r["name"]]
            matched.append((r, jn, "concept", why))
            continue
        if r["name"] in G.ARGUABLE_MATCHES:
            jn, why = G.ARGUABLE_MATCHES[r["name"]]
            arguable.append((r, jn, why))
            continue
        only_ours.append(r)

    def our_row(r, jnames, basis, why):
        name = r["name"]
        udb, udb_rel = "", ""
        if r.get("flag"):
            # the one confirmed parameter that overlaps an existing UDB param
            udb, udb_rel = "MISALIGNED_MAX_ATOMICITY_GRANULE_SIZE", \
                "flagged-overlap"
        bits = [why] if why else []
        if name in G.TYPE_CONFLICTS:
            bits.append(G.TYPE_CONFLICTS[name])
        if name in G.ADJACENT:
            bits.append("Nearest entry of his: " + G.ADJACENT[name])
        if r.get("flag"):
            bits.append(r["flag"])
        if r.get("resolved_note"):
            bits.append("Resolved in review: " + r["resolved_note"])
        if not r["tags"]:
            bits.append("No `[#norm:]` anchor in the manual for this text.")
        elif not r["rules"]:
            bits.append(f"Anchor `{r['tags'][0]}` has no normative rule entry.")
        elif not r["impl_def_marked"]:
            bits.append(f"Rule for `{r['tags'][0]}` is not marked "
                        "`impl-def-behavior: true` in the manual.")
        bits += j_notes(jnames)
        return {
            **j_raw(jnames),
            "concept": CONCEPTS.get(name, ""),
            "on_lists": "both" if jnames else "ishaan",
            "udb_param": udb,
            "udb_relationship": udb_rel,
            "james_name": ", ".join(jnames),
            "james_domain": j_cell(jnames, "domain"),
            "james_source": j_cell(jnames, "source"),
            "ishaan_name": name,
            "ishaan_domain": r["domain"],
            "ishaan_source": f"{r['tags'][0] if r['tags'] else '(untagged)'} "
                             f"{r['chapter']}:{r['line']}",
            "ishaan_class": r["class"],
            "match_basis": basis,
            "review_status": r["status"],
            "confidence": "",
            "notes_errors": notes_for(bits),
        }

    for r, jn, basis, why in matched:
        rows.append(our_row(r, jn, basis, why))
    for r, jn, why in arguable:
        row = our_row(r, jn, "arguable", why)
        row["on_lists"] = "both (arguable)"
        rows.append(row)
    for r in only_ours:
        rows.append(our_row(r, [], "n/a", ""))

    # ---------------------------------------------------------- his-only ---
    ours_tags = {r["tags"][0] for r in ours if r["tags"]}
    claimed = {n for _, jn, _, _ in matched for n in jn}
    claimed |= {n for _, jn, _ in arguable for n in jn}
    only_james = [r for r in james
                  if not (set(r["tags"]) & ours_tags) and r["name"] not in claimed]

    dropped_in_udb = 0
    for e in sorted(only_james, key=lambda x: x["name"]):
        v = verdicts.get(e["name"])
        verdict = v["verdict"] if v else "UNSEGMENTED"
        if verdict == "in_udb":
            dropped_in_udb += 1
            continue

        # Prefer his own description; fall back to the adjudication reason
        # only when it actually opens with a description of the parameter.
        reason = v["reason"] if v else ""
        desc = JAMES_CONCEPTS.get(e["name"]) or \
            clean_description(raw.get(e["name"], {}).get("description"))
        if desc:
            concept, udb_note = first_sentence(desc)[0], reason
        else:
            lead, rest = first_sentence(reason)
            if lead and not COMPARISON_LEAD.match(lead):
                concept, udb_note = lead, rest
            else:
                concept, udb_note = synth_concept(e), reason

        bits = []
        if udb_note:
            bits.append("UDB check: " + udb_note)
        if verdict == "out_of_scope":
            bits.insert(0, "OUT OF SCOPE under our criteria: platform "
                           "integration or no architectural units. Kept for "
                           "the mentor to rule on.")
        elif verdict == "internal_dup":
            bits.insert(0, "Duplicated within his own list.")
        elif verdict == "UNSEGMENTED":
            bits.insert(0, "Not segmented against UDB.")
        bits += j_notes([e["name"]])

        rows.append({
            **j_raw([e["name"]]),
            "concept": concept,
            "on_lists": "james",
            "udb_param": ", ".join(v["udb"]) if v and v.get("udb") else "",
            "udb_relationship": "partial" if verdict == "in_udb_partial" else "",
            "james_name": e["name"],
            "james_domain": e["domain"],
            "james_source": "; ".join(e["impl_defs"]) or "(none)",
            "ishaan_name": "",
            "ishaan_domain": "",
            "ishaan_source": "",
            "ishaan_class": "",
            "match_basis": "n/a",
            "review_status": "unreviewed (his individual draft)",
            "confidence": "",
            "notes_errors": notes_for(bits),
        })

    for i, r in enumerate(rows, 1):
        r["id"] = f"P{i:03d}"

    # Defects on entries this list excludes (UDB already defines them). Held
    # so the exclusion does not silently lose a defect worth telling him.
    landed = {n for r in rows
              for n in (r["james_name"].split(", ") if r["james_name"] else [])}
    off_list = {k: v for k, v in DEFECTS.items() if k not in landed}
    return rows, dropped_in_udb, off_list


def write_csv(rows):
    with OUT_CSV.open("w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=COLUMNS)
        w.writeheader()
        for r in rows:
            w.writerow({c: scrub(r.get(c, "")) for c in COLUMNS})


def write_xlsx(rows):
    """The circulating copy: frozen header, filters, confidence ready to fill."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "master_list"
    ws.append(COLUMNS)

    head = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="44546A")
    conf_fill = PatternFill("solid", fgColor="FFF2CC")
    conf_i = COLUMNS.index("confidence")

    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font, cell.fill = head, head_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for r in rows:
        ws.append([scrub(r.get(c, "")) for c in COLUMNS])

    widths = {
        "id": 7, "concept": 52, "on_lists": 15, "udb_param": 30,
        "udb_relationship": 16, "james_name": 26, "james_domain": 26,
        "james_source": 26, "ishaan_name": 30, "ishaan_domain": 28,
        "ishaan_source": 34, "ishaan_class": 15, "match_basis": 12,
        "review_status": 28, "confidence": 14, "notes_errors": 70,
        "james_description": 60, "james_long_name": 28, "james_note": 40,
    }
    for i, c in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 18)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[conf_i].fill = conf_fill
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}")

    notes = wb.create_sheet("how_to_use")
    for line in [
        ["Master parameter list"],
        [],
        ["Scope", "The union of James Ball's list and ours, minus the "
         "entries UDB already defines outright."],
        ["Rows", "One per distinct parameter concept, keyed by concept and "
         "not by name -- exactly one of our 38 names matches one of his."],
        ["Partial overlaps", "Kept, with the UDB parameter in udb_param. A "
         "granularity or aspect mismatch is not coverage."],
        [],
        ["confidence", f"FILL THIS IN. Scale: {CONFIDENCE_SCALE}. How "
         "strongly you believe the row is a real, certifiable parameter. "
         "Add your initials next to the value if more than one person "
         "rates a row."],
        [],
        ["james_* columns", "Carried verbatim from his YAML. james_domain "
         "is the one normalised value, kept for the like-for-like comparison "
         "against ishaan_domain; james_type / width / range / array are his "
         "originals beside it."],
        ["' | ' in a cell", "The row cites more than one of his entries, "
         "because some of our parameters are umbrellas over several of his."],
        [],
        ["Standing", "Every row of ours was reviewed individually by Allen "
         "Baum and Umer. His are an initial individual effort, explicitly "
         "not Parameter SIG consensus. See review_status per row."],
        ["Generated", "Do not hand-edit except in the confidence column; "
         "regenerate with scripts/generate_master_list.py."],
    ]:
        notes.append(line)
    notes.column_dimensions["A"].width = 22
    notes.column_dimensions["B"].width = 100
    for row in notes.iter_rows():
        row[0].font = Font(bold=True)
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(OUT_XLSX)


def write_adoc(rows, dropped, off_list):
    both = sum(1 for r in rows if r["on_lists"].startswith("both"))
    ish = sum(1 for r in rows if r["on_lists"] == "ishaan")
    jam = sum(1 for r in rows if r["on_lists"] == "james")
    L = [G.HDR.replace("generate_asciidoc.py", "generate_master_list.py"),
         "= Master Parameter List", ":toc:", ":toclevels: 2", "",
         "One row per distinct parameter concept, unioned across James Ball's",
         "list and ours. Entries UDB already defines outright are excluded;",
         f"{dropped} of his entries were dropped on that basis.",
         "",
         "Entries that only *partially* overlap a UDB parameter are kept, with",
         "the overlap recorded in the UDB column. A partial overlap is a",
         "granularity or aspect mismatch, not coverage.",
         "",
         f"*{len(rows)} rows*: {both} on both lists, {ish} only ours, "
         f"{jam} only his.",
         "",
         "NOTE: The `confidence` column is intentionally empty. It is for",
         f"reviewers to fill in, on the scale `{CONFIDENCE_SCALE}`, recording",
         "how strongly they believe the row is a real, certifiable parameter.",
         "Ratings from Allen, Umer and James are all welcome in the same",
         "column; put initials alongside the value if more than one person",
         "rates a row.",
         "",
         "CAUTION: The two lists do not have equal standing. Every row of ours",
         "was reviewed individually by Allen Baum and Umer; his are an initial",
         "individual effort, explicitly not Parameter SIG consensus. The",
         "`review_status` column carries this per row.",
         "",
         "The table below shows the comparison columns only. The full row is",
         "in `master_list.csv` and `master_list.xlsx` alongside this file,",
         "which additionally carry **every field of James' schema verbatim**",
         "-- `long-name`, `description`, `type`, `width`, `range`, `array`,",
         "`note`, `func-of-field-name`, chapter, section and granularity --",
         "as their own `james_*` columns rather than folded into a derived",
         "value. `james_domain` is the one normalised column, kept for the",
         "like-for-like comparison against `ishaan_domain`.",
         "",
         "A cell containing ` | ` means the row cites more than one of his",
         "entries, because some of our parameters are umbrellas over several",
         "of his.",
         "",
         '[cols="6,20,8,14,12,16,14,10,12",options="header"]',
         "|===",
         "| ID | Concept | Lists | UDB param | James' name | James' domain "
         "| Our name | Our domain | Confidence"]
    for r in rows:
        L.append(
            f"| {r['id']} | {G.esc(r['concept'])} | {G.esc(r['on_lists'])} "
            f"| {G.esc(r['udb_param']) or '--'} "
            f"| {G.esc(r['james_name']) or '--'} "
            f"| {G.esc(r['james_domain']) or '--'} "
            f"| {G.esc(r['ishaan_name']) or '--'} "
            f"| {G.esc(r['ishaan_domain']) or '--'} | ")
    L += ["|===", "",
          "== Rows carrying a defect or an open question", ""]
    flagged = [r for r in rows if r["notes_errors"]]
    L.append(f"{len(flagged)} of {len(rows)} rows carry a note. Sources, "
             "criteria flags and defects are in the CSV; the substantive ones "
             "are reproduced here.")
    L.append("")
    for r in flagged:
        if "DEFECT" in r["notes_errors"] or "OPEN" in r["notes_errors"] \
                or "OUT OF SCOPE" in r["notes_errors"]:
            label = r["james_name"] or r["ishaan_name"]
            L.append(f"* `{label}` ({r['id']}) -- {G.esc(r['notes_errors'])}")
    L.append("")

    if off_list:
        L += ["== Defects on entries this list excludes", "",
              "These sit on entries of his that UDB already defines, so they",
              "have no row above. Recorded here so excluding the row does not",
              "lose the defect -- each is still worth passing back to him.",
              ""]
        for k, v in sorted(off_list.items()):
            L.append(f"* `{k}` -- {G.esc(v)}")
        L.append("")
    return "\n".join(L)


def main() -> int:
    rows, dropped, off_list = build()
    G.OUTDIR.mkdir(parents=True, exist_ok=True)
    write_csv(rows)
    write_xlsx(rows)
    OUT_ADOC.write_text(write_adoc(rows, dropped, off_list))

    both = sum(1 for r in rows if r["on_lists"].startswith("both"))
    ish = sum(1 for r in rows if r["on_lists"] == "ishaan")
    jam = sum(1 for r in rows if r["on_lists"] == "james")
    print(f"rows            : {len(rows)}")
    print(f"  on both lists : {both}")
    print(f"  only ours     : {ish}")
    print(f"  only his      : {jam}")
    print(f"dropped (in UDB): {dropped}")
    if off_list:
        print(f"defects on excluded entries, listed separately: "
              f"{sorted(off_list)}")
    print(f"wrote  : {OUT_CSV.relative_to(G.REPO)}")
    print(f"wrote  : {OUT_XLSX.relative_to(G.REPO)}")
    print(f"wrote  : {OUT_ADOC.relative_to(G.REPO)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
